import errno
import openpyxl
import os

from django.conf import settings
from django.contrib.staticfiles import finders
from status.services.get_team_places_service import GetTeamPlacesService, GetTeamsService


class GetStatusTableFileUrlService:
	template_path = 'main/data-template.xlsx'
	result_path = 'downloads/data.xlsx'

	@staticmethod
	def _get_template_file():
		template = finders.find(GetStatusTableFileUrlService.template_path)
		if template is None:
			raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), GetStatusTableFileUrlService.template_path)
		
		return template
	
	@staticmethod
	def _create_status_table_file(time_type):
		wb_obj = openpyxl.load_workbook(GetStatusTableFileUrlService._get_template_file())
		sheet_obj = wb_obj.active

		places = GetTeamPlacesService.execute(time_type)
		for i in range(2, len(places) + 2):
			sheet_obj.cell(row=i, column=1).value = places[i - 2]['place']
			sheet_obj.cell(row=i, column=2).value = places[i - 2]['name']
			sheet_obj.cell(row=i, column=3).value = places[i - 2]['time']

		wb_obj.save(settings.MEDIA_ROOT / GetStatusTableFileUrlService.result_path)

	@staticmethod
	def execute(time_type):
		GetStatusTableFileUrlService._create_status_table_file(time_type)
		return settings.MEDIA_URL + GetStatusTableFileUrlService.result_path
